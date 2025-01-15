import json
import openpyxl


class DataFile:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = openpyxl.load_workbook(self.filename)
        self.sheet = self.workbook.active

    def fetch_row_count(self):
        return self.sheet.max_row

    def fetch_column_count(self):
        return self.sheet.max_column

    def fetch_keylist(self):
        col = self.fetch_column_count()
        keylist = []
        for key in range(1, col + 1):
            cel = self.sheet.cell(row=1, column=key)
            keylist.append(cel.value)

        return keylist

    def update_request_with_data(self, rownumber, keylist):
        data = {}
        col = self.fetch_column_count()
        for i in range(2, col + 1):
            cell = self.sheet.cell(row=rownumber, column=i)
            value = str(cell.value)
            data[keylist[i - 2]] = value

        return data
